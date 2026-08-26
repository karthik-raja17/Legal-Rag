"""
Production-grade ingestion job for Google Drive Excel + PDFs.

Features:
- Reads ALL sheets that contain a PDF link column.
- Extracts hyperlink targets correctly.
- Skips blank rows, rows without a valid link, and rows with non-URL placeholder text.
- Always stages PDFs through GCS and calls /parse_from_gcs (avoids 413s and removes
  the fragile file-size branch that broke when Drive metadata failed).
- Concurrent processing with configurable parallelism.
- Idempotent (skips already indexed docs, and remembers permanently-missing files
  so they aren't retried on every run).
- Tracks status in Firestore.
"""
import asyncio
import logging
import os
import re
import tempfile
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional, Dict, Any, List

import aiohttp
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from src.adapters.drive_client import DriveClient, DriveFileNotFound
from src.adapters.firestore_client import FirestoreClient
from src.adapters.gcs_client import GCSClient
from src.config.settings import settings

logger = logging.getLogger(__name__)

# ----------------------------------------------------------------------------
# Configuration
# ----------------------------------------------------------------------------

# Cell values that are not Drive links - contract status notes, placeholders, etc.
# Matched case-insensitively against the full stripped cell text.
INVALID_LINK_TEXTS = {
    "cmad", "nda", "bail non signe", "bail non signé", "x", "n/a", "na",
    "-", "", "en cours", "a venir", "à venir", "tbd",
}


@dataclass
class IngestionConfig:
    excel_file_id: Optional[str] = None
    excel_local_path: Optional[str] = None
    parser_url: str = os.environ.get("PARSER_URL", "https://your-parser-service-url")
    firestore_collection: str = settings.FIRESTORE_COLLECTION
    concurrency: int = 5
    dry_run: bool = False
    skip_existing: bool = True
    # Column names that contain the PDF link in each sheet
    pdf_link_columns: Dict[str, str] = field(default_factory=lambda: {
        "CLIENTS": "Lien vers le contrat",
        "IMMOBILIER": "Lien vers le contrat",
        "ACHAT CENTRALE": "Lien vers le contrat",
        "EPC - FOURNISSEURS PV": "Lien vers le contrat",
        "FOURNISSEURS BATTERIE & IRVE": "Lien vers le contrat",
    })
    # Column for document ID (optional)
    doc_id_columns: Dict[str, str] = field(default_factory=lambda: {
        "CLIENTS": "Nom du projet",
        "IMMOBILIER": "Nom du Projet",
        "ACHAT CENTRALE": "Adresse site",  # fallback to address
        "EPC - FOURNISSEURS PV": "Nom du Projet",
        "FOURNISSEURS BATTERIE & IRVE": "Nom du Projet",
    })


# ----------------------------------------------------------------------------
# Ingestion Engine
# ----------------------------------------------------------------------------
class IngestionEngine:
    def __init__(self, config: IngestionConfig):
        self.config = config
        self.drive = DriveClient()
        self.firestore = FirestoreClient()
        self.gcs = GCSClient()
        self._shutdown = False
        # Track counts for a summary at the end of the run
        self._stats = {"processed": 0, "skipped_invalid_link": 0, "skipped_missing_drive": 0,
                        "skipped_already_done": 0, "failed": 0}

    def extract_drive_file_id(self, link: str) -> Optional[str]:
        """Extract file ID from a Google Drive link or raw ID. Returns None for
        anything that isn't plausibly a Drive link/ID (status notes, blanks, etc.)."""
        if not link or not isinstance(link, str):
            return None
        link = link.strip()
        if not link:
            return None
        if link.lower() in INVALID_LINK_TEXTS:
            return None
        # A plain Drive file ID is a long alphanumeric/-/_ token (typically 25-44 chars)
        # and contains no spaces. Short tokens like "x" are excluded via the set above
        # and via the length check below to avoid false positives.
        if re.match(r"^[a-zA-Z0-9_-]{15,}$", link):
            return link
        # Standard share links: /d/ID/
        match = re.search(r"/d/([a-zA-Z0-9_-]+)", link)
        if match:
            return match.group(1)
        # ?id=... style
        from urllib.parse import urlparse, parse_qs
        parsed = urlparse(link)
        qs = parse_qs(parsed.query)
        file_id = qs.get("id", [None])[0]
        if file_id:
            return file_id
        # Doesn't look like a link at all (e.g. free-text note) - not an error, just skip.
        return None

    # In src/orchestrator/ingestor.py, inside class IngestionEngine, after `_call_parse_document`:

    async def _call_parse_from_gcs(self, document_id: str, gcs_uri: str):
        url = f"{self.config.parser_url}/parse_from_gcs"
        data = {"document_id": document_id, "gcs_uri": gcs_uri}
        async with aiohttp.ClientSession() as session:
            async with session.post(url, data=data, timeout=300) as resp:
                resp.raise_for_status()
                return await resp.json()

    def get_cell_value_or_hyperlink(self, cell: Cell) -> Optional[str]:
        """Extract the hyperlink target if present, otherwise the cell value."""
        if cell.hyperlink and cell.hyperlink.target:
            return cell.hyperlink.target
        val = cell.value
        if isinstance(val, str) and (val.startswith("http") or val.startswith("https")):
            return val
        return val

    async def process_row(
        self,
        sheet_name: str,
        headers: List[str],
        row: List[Cell],
        row_idx: int,
        semaphore: asyncio.Semaphore,
    ):
        async with semaphore:
            await self._process_row_internal(sheet_name, headers, row, row_idx)

    async def _process_row_internal(self, sheet_name: str, headers: List[str], row: List[Cell], row_idx: int):
        pdf_col_name = self.config.pdf_link_columns.get(sheet_name)
        if not pdf_col_name:
            logger.warning(f"Sheet {sheet_name} has no PDF link column mapping, skipping.")
            return

        try:
            col_idx = headers.index(pdf_col_name)
        except ValueError:
            logger.warning(f"Sheet {sheet_name} missing column '{pdf_col_name}', skipping.")
            return

        if col_idx >= len(row):
            return

        cell = row[col_idx]
        link = self.get_cell_value_or_hyperlink(cell)

        drive_file_id = self.extract_drive_file_id(link) if link else None
        if not drive_file_id:
            logger.info(f"Row {row_idx} in {sheet_name}: no valid Drive link ('{link}'), skipping.")
            self._stats["skipped_invalid_link"] += 1
            return

        # Generate document ID up front so we can check/write Firestore state
        # even before hitting Drive.
        doc_id = self._resolve_doc_id(sheet_name, headers, row, row_idx)

        # Check if already processed or permanently unavailable - idempotency,
        # and avoids re-hitting Drive for files we already know are 404/403.
        if self.config.skip_existing:
            state = self.firestore.get_document_state(doc_id)
            if state and state.get("status") in ("indexed", "processing", "missing_in_drive"):
                logger.info(f"Row {row_idx} in {sheet_name}: {doc_id} already handled "
                            f"(status={state.get('status')}), skipping.")
                self._stats["skipped_already_done"] += 1
                return

        # Get metadata from Drive - fails fast (no retries) on 404/403.
        try:
            meta = self.drive.get_file_metadata(drive_file_id)
        except DriveFileNotFound as e:
            logger.warning(f"Row {row_idx} in {sheet_name}: {drive_file_id} not found/accessible "
                            f"in Drive (HTTP {e.status}), recording and skipping.")
            self.firestore.update_document_state(doc_id, {
                "status": "missing_in_drive",
                "drive_file_id": drive_file_id,
                "source_sheet": sheet_name,
                "source_row": row_idx,
                "http_status": e.status,
            })
            self._stats["skipped_missing_drive"] += 1
            return
        except Exception as e:
            logger.error(f"Row {row_idx} in {sheet_name}: failed to get metadata for {drive_file_id}: {e}")
            self._stats["failed"] += 1
            return

        file_size = int(meta.get("size", 0) or 0)
        modified_time = meta.get("modifiedTime")
        file_name = meta.get("name", f"row_{row_idx}")

        if self.config.dry_run:
            logger.info(f"DRY RUN: would process {doc_id} ({file_name}, {file_size/1024/1024:.1f} MB)")
            return

        # Download PDF
        try:
            logger.info(f"Row {row_idx} in {sheet_name}: downloading {drive_file_id} ({file_size/1024/1024:.1f} MB)")
            pdf_content = self.drive.download_file(drive_file_id)
        except DriveFileNotFound as e:
            logger.warning(f"Row {row_idx} in {sheet_name}: {drive_file_id} disappeared between "
                            f"metadata and download (HTTP {e.status}), recording and skipping.")
            self.firestore.update_document_state(doc_id, {
                "status": "missing_in_drive",
                "drive_file_id": drive_file_id,
                "source_sheet": sheet_name,
                "source_row": row_idx,
                "http_status": e.status,
            })
            self._stats["skipped_missing_drive"] += 1
            return
        except Exception as e:
            logger.error(f"Row {row_idx} in {sheet_name}: failed to download {drive_file_id}: {e}")
            self.firestore.update_document_state(doc_id, {
                "status": "failed",
                "error": str(e),
                "drive_file_id": drive_file_id,
                "source_sheet": sheet_name,
            })
            self._stats["failed"] += 1
            return

        # Always stage in GCS - this is now mandatory, not just for large files.
        # Removing the size-based branch also removes the bug where a failed
        # metadata call silently defaulted file_size to 0 and routed large PDFs
        # through the direct-upload path, causing 413s.
        try:
            blob_name = f"pdfs/{drive_file_id}.pdf"
            self.gcs.upload_pdf(drive_file_id, pdf_content, modified_time)
            gcs_uri = f"gs://{self.gcs.bucket_name}/{blob_name}"
        except Exception as e:
            logger.error(f"Row {row_idx} in {sheet_name}: GCS upload failed for {drive_file_id}: {e}")
            self.firestore.update_document_state(doc_id, {
                "status": "failed",
                "error": f"GCS upload failed: {e}",
                "drive_file_id": drive_file_id,
                "source_sheet": sheet_name,
            })
            self._stats["failed"] += 1
            return

        # Trigger parsing - always via /parse_from_gcs now.
        try:
            await self._call_parse_from_gcs(doc_id, gcs_uri)
            self.firestore.update_document_state(doc_id, {
                "status": "processing",
                "drive_file_id": drive_file_id,
                "source_sheet": sheet_name,
                "source_row": row_idx,
                "gcs_uri": gcs_uri,
            })
            logger.info(f"Row {row_idx} in {sheet_name}: parsing triggered for {doc_id}")
            self._stats["processed"] += 1
        except Exception as e:
            logger.error(f"Row {row_idx} in {sheet_name}: failed to trigger parsing: {e}")
            self.firestore.update_document_state(doc_id, {
                "status": "failed",
                "error": str(e),
                "drive_file_id": drive_file_id,
                "source_sheet": sheet_name,
            })
            self._stats["failed"] += 1

    def _resolve_doc_id(self, sheet_name: str, headers: List[str], row: List[Cell], row_idx: int) -> str:
        doc_id = None
        doc_col_name = self.config.doc_id_columns.get(sheet_name)
        if doc_col_name:
            try:
                doc_idx = headers.index(doc_col_name)
                if doc_idx < len(row):
                    doc_cell = row[doc_idx]
                    doc_id = doc_cell.value if doc_cell.value else None
            except ValueError:
                pass
        if not doc_id:
            doc_id = f"{sheet_name}_{row_idx}"
        return re.sub(r"[^a-zA-Z0-9_\-]", "_", str(doc_id))[:64]

    async def _call_parse_document(self, document_id: str, pdf_content: bytes, filename: str):
        """
        Calls the current parser API's /parse endpoint (multipart file upload),
        replacing the old /parse_from_gcs (GCS-URI reference) contract that the
        parser no longer exposes.
        """
        url = f"{self.config.parser_url}/parse"
        form = aiohttp.FormData()
        form.add_field(
            "file",
            pdf_content,
            filename=filename,
            content_type="application/pdf",
        )
        form.add_field("document_id", document_id)

        async with aiohttp.ClientSession() as session:
            async with session.post(url, data=form, timeout=aiohttp.ClientTimeout(total=120)) as resp:
                resp.raise_for_status()
                return await resp.json()

    async def run(self):
        # Download Excel to a temporary file
        if self.config.excel_local_path:
            excel_path = self.config.excel_local_path
        else:
            if not self.config.excel_file_id:
                raise ValueError("No excel_file_id or excel_local_path provided")
            excel_content = self.drive.download_file(self.config.excel_file_id)
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp.write(excel_content)
                excel_path = tmp.name

        wb = load_workbook(excel_path, data_only=False)
        semaphore = asyncio.Semaphore(self.config.concurrency)

        sheets_to_process = [
            "CLIENTS",
            "IMMOBILIER",
            "ACHAT CENTRALE",
            "EPC - FOURNISSEURS PV",
            "FOURNISSEURS BATTERIE & IRVE",
        ]

        for sheet_name in sheets_to_process:
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet {sheet_name} not found, skipping.")
                continue

            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=False))
            if not rows:
                continue
            headers_cells = rows[0]
            headers = [cell.value for cell in headers_cells]
            pdf_col_name = self.config.pdf_link_columns.get(sheet_name)
            if not pdf_col_name:
                continue
            try:
                pdf_idx = headers.index(pdf_col_name)
            except ValueError:
                logger.warning(f"Sheet {sheet_name}: column '{pdf_col_name}' not found, skipping.")
                continue

            logger.info(f"Processing sheet {sheet_name} with {len(rows)-1} data rows.")

            tasks = []
            for idx, row in enumerate(rows[1:], start=1):
                if pdf_idx >= len(row):
                    continue
                cell = row[pdf_idx]
                link = self.get_cell_value_or_hyperlink(cell)
                if not link:
                    continue
                tasks.append(self.process_row(sheet_name, headers, row, idx, semaphore))

            await asyncio.gather(*tasks, return_exceptions=True)

        if not self.config.excel_local_path and os.path.exists(excel_path):
            os.unlink(excel_path)

        logger.info(
            "Ingestion job completed. "
            f"processed={self._stats['processed']} "
            f"skipped_invalid_link={self._stats['skipped_invalid_link']} "
            f"skipped_missing_drive={self._stats['skipped_missing_drive']} "
            f"skipped_already_done={self._stats['skipped_already_done']} "
            f"failed={self._stats['failed']}"
        )

    def shutdown(self):
        self._shutdown = True


# ----------------------------------------------------------------------------
# Entrypoint
# ----------------------------------------------------------------------------
async def main_async():
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel-file-id", default=os.environ.get("EXCEL_FILE_ID"))
    parser.add_argument("--excel-path")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--concurrency", type=int, default=5)
    args = parser.parse_args()

    logging.basicConfig(level=logging.INFO)

    config = IngestionConfig(
        excel_file_id=args.excel_file_id,
        excel_local_path=args.excel_path,
        dry_run=args.dry_run,
        concurrency=args.concurrency,
    )
    engine = IngestionEngine(config)
    await engine.run()


def main():
    asyncio.run(main_async())


if __name__ == "__main__":
    main()